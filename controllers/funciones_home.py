
# Para subir archivo tipo foto al servidor
from werkzeug.utils import secure_filename
import uuid  # Modulo de python para crear un string

from conexion.conexionBD import connectionBD  # Conexión a BD

import datetime
import re
import os

from os import remove  # Modulo  para remover archivo
from os import path  # Modulo para obtener la ruta o directorio


import openpyxl  # Para generar el excel
# biblioteca o modulo send_file para forzar la descarga
from flask import send_file

#CLIENTES
def procesar_form_cliente(dataForm):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO tbl_clientes (nombre_cliente, apellido_cliente, nit_cliente, direccion_cliente) VALUES (%s, %s, %s, %s)"

                # Creando una tupla con los valores del INSERT
                valores = (dataForm['nombre_cliente'], dataForm['apellido_cliente'], dataForm['nit_cliente'],
                           dataForm['direccion_cliente'])
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_cliente: {str(e)}'


# Lista de Clientes
def sql_lista_clientesBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT c.id_cliente,c.nombre_cliente,c.apellido_cliente,c.nit_cliente,c.direccion_cliente FROM tbl_clientes AS c ORDER BY c.id_cliente DESC"
                cursor.execute(querySQL)
                clientesBD = cursor.fetchall()
        return clientesBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_clientesBD: {e}")
        return None


# Detalles del Cliente
def sql_detalles_clientesBD(idCliente):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        c.id_cliente,
                        c.nombre_cliente, 
                        c.apellido_cliente,
                        c.nit_cliente,
                        c.direccion_cliente 
                    FROM tbl_clientes AS c
                    WHERE id_cliente =%s
                    ORDER BY c.id_cliente DESC
                    """)
                cursor.execute(querySQL, (idCliente,))
                clientesBD = cursor.fetchone()
        return clientesBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_clientesBD: {e}")
        return None
    

def buscarClienteBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            c.id_cliente,
                            c.nombre_cliente, 
                            c.apellido_cliente,
                            c.nit_cliente,
                            c.direccion_cliente
                        FROM tbl_clientes AS c
                        WHERE c.nombre_cliente LIKE %s 
                        ORDER BY c.id_cliente DESC
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def buscarClienteBD: {e}")
        return []


def buscarClienteUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            c.id_cliente,
                            c.nombre_cliente, 
                            c.apellido_cliente,
                            c.nit_cliente,
                            c.direccion_cliente
                        FROM tbl_clientes AS c
                        WHERE c.id_cliente =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                cliente = mycursor.fetchone()
                return cliente

    except Exception as e:
        print(f"Ocurrió un error en def buscarClienteUnico: {e}")
        return []


def procesar_actualizacion_formCliente(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                nombre_cliente = data.form['nombre_cliente']
                apellido_cliente = data.form['apellido_cliente']
                nit_cliente = data.form['nit_cliente']
                direccion_cliente = data.form['direccion_cliente']
                id_cliente = data.form['id_cliente']
                querySQL = """
                        UPDATE tbl_clientes
                        SET 
                            nombre_cliente = %s,
                            apellido_cliente= %s,
                            nit_cliente = %s,
                            direccion_cliente = %s
                        WHERE id_cliente = %s
                    """
                values = (nombre_cliente, apellido_cliente, nit_cliente,direccion_cliente,id_cliente)
                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_formCliente: {e}")
        return None
# Eliminar cliente
def eliminarCliente(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM tbl_clientes WHERE id_cliente=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarCliente: {e}")
        return []

#PROVEEDORES
def procesar_form_proveedor(dataForm):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO tbl_proveedores (nombre_proveedor, nit_proveedor, direccion_proveedor, telefono_proveedor) VALUES (%s, %s, %s, %s)"
                # Creando una tupla con los valores del INSERT
                valores = (dataForm['nombre_proveedor'],dataForm['nit_proveedor'],
                           dataForm['direccion_proveedor'],dataForm['telefono_proveedor'])
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_proveedor: {str(e)}'

# Lista de Proveedores
def sql_lista_proveedoresBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT p.id_proveedor,p.nombre_proveedor,p.nit_proveedor,p.direccion_proveedor,p.telefono_proveedor FROM tbl_proveedores AS p ORDER BY p.id_proveedor DESC"
                cursor.execute(querySQL)
                proveedoresBD = cursor.fetchall()
        return proveedoresBD
    except Exception as e:
        print(
            f"Error en la función sql_lista_proveedoresBD: {e}")
        return None

# Detalles del Proveedor
def sql_detalles_proveedoresBD(idProveedor):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        p.id_proveedor,
                        p.nombre_proveedor, 
                        p.nit_proveedor,
                        p.direccion_proveedor,
                        p.telefono_proveedor 
                    FROM tbl_proveedores AS p
                    WHERE id_proveedor =%s
                    ORDER BY p.id_proveedor DESC
                    """)
                cursor.execute(querySQL, (idProveedor,))
                proveedorBD = cursor.fetchone()
        return proveedorBD
    except Exception as e:
        print(
            f"Error en la función sql_detalles_proveedoresBD: {e}")
        return None
    

def buscarProveedorBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            p.id_proveedor,
                            p.nombre_proveedor, 
                            p.nit_proveedor,
                            p.direccion_proveedor,
                            p.telefono_proveedor
                        FROM tbl_proveedores AS p
                        WHERE p.nombre_proveedor LIKE %s 
                        ORDER BY p.id_proveedor DESC
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def buscarProveedorBD: {e}")
        return []


def buscarProveedorUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            p.id_proveedor,
                            p.nombre_proveedor,
                            p.nit_proveedor,
                            p.direccion_proveedor,
                            p.telefono_proveedor
                        FROM tbl_proveedores AS p
                        WHERE p.id_proveedor =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                proveedor = mycursor.fetchone()
                return proveedor

    except Exception as e:
        print(f"Ocurrió un error en def buscarProveedorUnico: {e}")
        return []


def procesar_actualizacion_formProveedor(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                nombre_proveedor = data.form['nombre_proveedor']
                nit_proveedor = data.form['nit_proveedor']
                direccion_proveedor = data.form['direccion_proveedor']
                telefono_proveedor = data.form['telefono_proveedor']
                id_proveedor = data.form['id_proveedor']
                querySQL = """
                        UPDATE tbl_proveedores
                        SET 
                            nombre_proveedor = %s,
                            nit_proveedor = %s,
                            direccion_proveedor = %s,
                            telefono_proveedor = %s
                        WHERE id_proveedor = %s
                    """
                values = (nombre_proveedor, nit_proveedor, direccion_proveedor, telefono_proveedor, id_proveedor)
                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_formProveedor: {e}")
        return None
# Eliminar proveedor
def eliminarProveedor(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM tbl_proveedores WHERE id_proveedor=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarProveedor: {e}")
        return []

#EMPLEADOS
def procesar_form_empleado(dataForm, foto_perfil):
    # Formateando Salario
    salario_sin_puntos = re.sub('[^0-9]+', '', dataForm['salario_empleado'])
    # convertir salario a INT
    salario_entero = int(salario_sin_puntos)

    result_foto_perfil = procesar_imagen_perfil(foto_perfil)
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO tbl_empleados (nombre_empleado, apellido_empleado, sexo_empleado, telefono_empleado, email_empleado, profesion_empleado, foto_empleado, salario_empleado) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"

                # Creando una tupla con los valores del INSERT
                valores = (dataForm['nombre_empleado'], dataForm['apellido_empleado'], dataForm['sexo_empleado'],
                           dataForm['telefono_empleado'], dataForm['email_empleado'], dataForm['profesion_empleado'], result_foto_perfil, salario_entero)
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_empleado: {str(e)}'


def procesar_imagen_perfil(foto):
    try:
        # Nombre original del archivo
        filename = secure_filename(foto.filename)
        extension = os.path.splitext(filename)[1]

        # Creando un string de 50 caracteres
        nuevoNameFile = (uuid.uuid4().hex + uuid.uuid4().hex)[:100]
        nombreFile = nuevoNameFile + extension

        # Construir la ruta completa de subida del archivo
        basepath = os.path.abspath(os.path.dirname(__file__))
        upload_dir = os.path.join(basepath, f'../static/fotos_empleados/')

        # Validar si existe la ruta y crearla si no existe
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            # Dando permiso a la carpeta
            os.chmod(upload_dir, 0o755)

        # Construir la ruta completa de subida del archivo
        upload_path = os.path.join(upload_dir, nombreFile)
        foto.save(upload_path)

        return nombreFile

    except Exception as e:
        print("Error al procesar archivo:", e)
        return []


# Lista de Empleados
def sql_lista_empleadosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = (f"""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.salario_empleado,
                        e.foto_empleado,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_empleadosBD: {e}")
        return None


# Detalles del Empleado
def sql_detalles_empleadosBD(idEmpleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.salario_empleado,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado,
                        e.telefono_empleado, 
                        e.email_empleado,
                        e.profesion_empleado,
                        e.foto_empleado,
                        DATE_FORMAT(e.fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro
                    FROM tbl_empleados AS e
                    WHERE id_empleado =%s
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL, (idEmpleado,))
                empleadosBD = cursor.fetchone()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_empleadosBD: {e}")
        return None


# Funcion Empleados Informe (Reporte)
def empleadosReporte():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.salario_empleado,
                        e.email_empleado,
                        e.telefono_empleado,
                        e.profesion_empleado,
                        DATE_FORMAT(e.fecha_registro, '%d de %b %Y %h:%i %p') AS fecha_registro,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función empleadosReporte: {e}")
        return None


def generarReporteExcel():
    dataEmpleados = empleadosReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Agregar la fila de encabezado con los títulos
    cabeceraExcel = ("Nombre", "Apellido", "Sexo",
                     "Telefono", "Email", "Profesión", "Salario", "Fecha de Ingreso")

    hoja.append(cabeceraExcel)

    # Formato para números en moneda colombiana y sin decimales
    formato_moneda_colombiana = '#,##0'

    # Agregar los registros a la hoja
    for registro in dataEmpleados:
        nombre_empleado = registro['nombre_empleado']
        apellido_empleado = registro['apellido_empleado']
        sexo_empleado = registro['sexo_empleado']
        telefono_empleado = registro['telefono_empleado']
        email_empleado = registro['email_empleado']
        profesion_empleado = registro['profesion_empleado']
        salario_empleado = registro['salario_empleado']
        fecha_registro = registro['fecha_registro']

        # Agregar los valores a la hoja
        hoja.append((nombre_empleado, apellido_empleado, sexo_empleado, telefono_empleado, email_empleado, profesion_empleado,
                     salario_empleado, fecha_registro))

        # Itera a través de las filas y aplica el formato a la columna G
        for fila_num in range(2, hoja.max_row + 1):
            columna = 7  # Columna G
            celda = hoja.cell(row=fila_num, column=columna)
            celda.number_format = formato_moneda_colombiana

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_empleados_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        # Dando permisos a la carpeta
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)


def buscarEmpleadoBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.nombre_empleado, 
                            e.apellido_empleado,
                            e.salario_empleado,
                            CASE
                                WHEN e.sexo_empleado = 1 THEN 'Masculino'
                                ELSE 'Femenino'
                            END AS sexo_empleado
                        FROM tbl_empleados AS e
                        WHERE e.nombre_empleado LIKE %s 
                        ORDER BY e.id_empleado DESC
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoBD: {e}")
        return []


def buscarEmpleadoUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.nombre_empleado, 
                            e.apellido_empleado,
                            e.sexo_empleado,
                            e.telefono_empleado,
                            e.email_empleado,
                            e.profesion_empleado,
                            e.salario_empleado,
                            e.foto_empleado
                        FROM tbl_empleados AS e
                        WHERE e.id_empleado =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                empleado = mycursor.fetchone()
                return empleado

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoUnico: {e}")
        return []


def procesar_actualizacion_form(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                nombre_empleado = data.form['nombre_empleado']
                apellido_empleado = data.form['apellido_empleado']
                sexo_empleado = data.form['sexo_empleado']
                telefono_empleado = data.form['telefono_empleado']
                email_empleado = data.form['email_empleado']
                profesion_empleado = data.form['profesion_empleado']

                salario_sin_puntos = re.sub(
                    '[^0-9]+', '', data.form['salario_empleado'])
                salario_empleado = int(salario_sin_puntos)
                id_empleado = data.form['id_empleado']

                if data.files['foto_empleado']:
                    file = data.files['foto_empleado']
                    fotoForm = procesar_imagen_perfil(file)

                    querySQL = """
                        UPDATE tbl_empleados
                        SET 
                            nombre_empleado = %s,
                            apellido_empleado = %s,
                            sexo_empleado = %s,
                            telefono_empleado = %s,
                            email_empleado = %s,
                            profesion_empleado = %s,
                            salario_empleado = %s,
                            foto_empleado = %s
                        WHERE id_empleado = %s
                    """
                    values = (nombre_empleado, apellido_empleado, sexo_empleado,
                              telefono_empleado, email_empleado, profesion_empleado,
                              salario_empleado, fotoForm, id_empleado)
                else:
                    querySQL = """
                        UPDATE tbl_empleados
                        SET 
                            nombre_empleado = %s,
                            apellido_empleado = %s,
                            sexo_empleado = %s,
                            telefono_empleado = %s,
                            email_empleado = %s,
                            profesion_empleado = %s,
                            salario_empleado = %s
                        WHERE id_empleado = %s
                    """
                    values = (nombre_empleado, apellido_empleado, sexo_empleado,
                              telefono_empleado, email_empleado, profesion_empleado,
                              salario_empleado, id_empleado)

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form: {e}")
        return None


# Lista de Usuarios creados
def lista_usuariosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT id, name_surname, email_user, created_user FROM users"
                cursor.execute(querySQL,)
                usuariosBD = cursor.fetchall()
        return usuariosBD
    except Exception as e:
        print(f"Error en lista_usuariosBD : {e}")
        return []


# Eliminar uEmpleado-------------------------------------------
def eliminarEmpleado(id_empleado, foto_empleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM tbl_empleados WHERE id_empleado=%s"
                cursor.execute(querySQL, (id_empleado,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

                if resultado_eliminar:
                    # Eliminadon foto_empleado desde el directorio
                    basepath = path.dirname(__file__)
                    url_File = path.join(
                        basepath, '../static/fotos_empleados', foto_empleado)

                    if path.exists(url_File):
                        remove(url_File)  # Borrar foto desde la carpeta

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarEmpleado : {e}")
        return []


# Eliminar usuario----------------------
def eliminarUsuario(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM users WHERE id=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarUsuario : {e}")
        return []

#PRODUCTOS-----------------------------------------------
def procesar_form_producto(dataForm):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO tbl_productos (nombre_producto, descripcion_producto, cantidad_producto, precio_producto, categoria_producto) VALUES (%s, %s, %s, %s, %s)"

                # Creando una tupla con los valores del INSERT
                valores = (dataForm['nombre_producto'], dataForm['descripcion_producto'], dataForm['cantidad_producto'],
                           dataForm['precio_producto'],dataForm['categoria_producto'])
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_producto: {str(e)}'


# Lista de Productos
def sql_lista_productosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT pr.id_producto,pr.nombre_producto,pr.descripcion_producto,pr.cantidad_producto,pr.precio_producto, CASE WHEN pr.categoria_producto = 1 THEN 'Alimentos' ELSE 'Higiene y salud' END AS categoria_producto FROM tbl_productos AS pr ORDER BY pr.id_producto DESC"
                cursor.execute(querySQL)
                clientesBD = cursor.fetchall()
        return clientesBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_productosBD: {e}")
        return None


# Detalles del Producto
def sql_detalles_productosBD(idProducto):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        pr.id_producto,
                        pr.nombre_producto, 
                        pr.descripcion_producto,
                        pr.cantidad_producto,
                        pr.precio_producto,
                        CASE
                            WHEN pr.categoria_producto = 1 THEN 'Alimientos'
                            ELSE 'Higiene y salud'
                        END AS categoria_producto 
                    FROM tbl_productos AS pr
                    WHERE id_producto =%s
                    ORDER BY pr.id_producto DESC
                    """)
                cursor.execute(querySQL, (idProducto,))
                productosBD = cursor.fetchone()
        return productosBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_productosBD: {e}")
        return None
    

def buscarProductoBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            pr.id_producto,
                            pr.nombre_producto,
                            pr.descripcion_producto, 
                            pr.cantidad_producto,
                            pr.precio_producto,
                            pr.categoria_producto
                        FROM tbl_productos AS pr
                        WHERE pr.nombre_producto LIKE %s 
                        ORDER BY pr.id_producto DESC
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def buscarProductoBD: {e}")
        return []


def buscarProductoUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            pr.id_producto,
                            pr.nombre_producto,
                            pr.descripcion_producto, 
                            pr.cantidad_producto,
                            pr.precio_producto,
                            pr.categoria_producto
                        FROM tbl_productos AS pr
                        WHERE pr.id_producto =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                cliente = mycursor.fetchone()
                return cliente

    except Exception as e:
        print(f"Ocurrió un error en def buscarProductoUnico: {e}")
        return []


def procesar_actualizacion_formProducto(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                nombre_producto = data.form['nombre_producto']
                descripcion_producto = data.form['descripcion_producto']
                cantidad_producto = data.form['cantidad_producto']
                precio_producto = data.form['precio_producto']
                categoria_producto = data.form['categoria_producto']
                id_producto = data.form['id_producto']
                querySQL = """
                        UPDATE tbl_productos
                        SET 
                            nombre_producto = %s,
                            descripcion_producto= %s,
                            cantidad_producto = %s,
                            precio_producto = %s,
                            categoria_producto = %s
                        WHERE id_producto = %s
                    """
                values = (nombre_producto, descripcion_producto, cantidad_producto,precio_producto, categoria_producto,id_producto)
                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_formProducto: {e}")
        return None
# Eliminar producto
def eliminarProducto(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM tbl_productos WHERE id_producto=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarProducto: {e}")
        return []

